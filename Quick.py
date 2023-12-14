import datetime
import requests
from bs4 import BeautifulSoup
import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime

#import schedule
import time

def get_all_pages():
    urls = []

    for page_number in range(1, 4):  
        if page_number == 1:
            i = "https://www.avis-clients.fr/site/avis-quick.fr"
        else:
            i = f"https://www.avis-clients.fr/site/avis-quick.fr_{page_number}"
        urls.append(i)

    return urls

def parse_comments(url):
    for url in urls:
        request = requests.get(url)
        print(request.status_code)
        soup = BeautifulSoup(request.content, "html.parser")
        reviews = soup.find_all('div', class_='review-single pt-30')
        for review in reviews:
            
            try:
                comment = review.find("p", class_="copy justify").text.strip()
            except AttributeError as e:
                comment = ""
            
            date_review = review.find("span", class_="review-holder-name h5 pb-10").text.strip()
            #print(date_review)

            likes = review.find("span", class_="rating-stars")["data-rating"]
            #print(likes)

            header = ['Enterprise', 'Comments', "Reviews_Date", "Evaluations", "Extraction-Date", "Category", "Source"]
            data = ["Quick", comment, date_review, likes, datetime.today().date(), "Fast food", "Avis-Clients.fr"]
            chemin = "D:/Lorraine/Fichiers/Quick-Review.xlsx"

            file_exists = openpyxl.workbook.Workbook()

            file_exists = os.path.isfile(chemin)

            # Création d'un classeur s'il n'existe pas et écriture des en-têtes
            if not file_exists:
                wb = Workbook()
                ws = wb.active
                ws.append(header)
                wb.save(chemin)

            # Écriture des données
            wb = openpyxl.load_workbook(chemin)
            ws = wb.active
            ws.append(data)
            wb.save(chemin)

urls = get_all_pages()
parse_comments(urls)