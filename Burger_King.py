import csv
import datetime
import requests
from bs4 import BeautifulSoup
import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime

#import schedule
import time

def get_all_pages():
    urls = []

    page_number = 1

    for i in range(26):
        i = f"https://fr.trustpilot.com/review/www.burgerking.fr?page={page_number}"
        page_number += 1
        urls.append(i)

    return urls

def parse_comments(url):
    for url in urls:
        request = requests.get(url)
        print(request.status_code)
        soup = BeautifulSoup(request.content, "html.parser")
        reviews = soup.find_all('div', class_='styles_reviewCardInner__EwDq2')
        for review in reviews:
            
            try:
                comment = review.find("p", class_="typography_body-l__KUYFJ typography_appearance-default__AAY17 typography_color-black__5LYEn").text.strip()
            except AttributeError as e:
                comment = ""
            
            date_review = review.find("time").get_text(strip=True)
            #print(date_review)

            likes = review.find("img")['alt']
        # print(likes)

            
            header = ['Enterprise', 'Comments', "Reviews_Date", "Evaluations", "Extraction-Date", "Category", "Source"]
            data = ["Burger King", comment, date_review, likes, datetime.today().date(),"Fast food", "Trustpilot"]

            chemin = "D:/Lorraine/Fichiers/Burger_King-Review.xlsx"

            file_exists = openpyxl.workbook.Workbook()

            file_exists = os.path.isfile(chemin)

            # Création d'un classeur s'il n'existe pas et écriture des en-têtes
            if not file_exists:
                wb = Workbook()
                ws = wb.active
                ws.append(header)
                wb.save(chemin)

            # Écriture des données
            wb = openpyxl.load_workbook(chemin)
            ws = wb.active
            ws.append(data)
            wb.save(chemin)


urls = get_all_pages()
parse_comments(urls)