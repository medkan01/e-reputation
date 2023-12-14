import datetime
import requests
from bs4 import BeautifulSoup
import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime

#import schedule
import time

def get_all_pages():
    base_url = "https://www.avis-verifies.com/avis-clients/lecoqsportif.com"
    urls = []

    for page_number in range(1, 27):
        if page_number == 1:
            url = base_url
        else:
            url = f"{base_url}?p=28{chr(96 + page_number)}"

        urls.append(url)

    return urls


def parse_comments(url):
    for url in urls:
        request = requests.get(url)
        print(request.status_code)
        soup = BeautifulSoup(request.content, "html.parser")
        reviews = soup.find_all('li', class_='reviews__item review')
        for review in reviews:
            try:
                comment = review.find("p", class_="review__text search-criterion").text.strip()
            except AttributeError as e:
                comment = ""
            
            date_review = review.find("time", class_="review__data-time").text.strip()
            #print(date_review)

            likes = review.find("span", class_="review__rating-fact").text.strip()
            # print(likes)

            
            header = ['Enterprise', 'Comments', "Reviews_Date", "Evaluations", "Extraction-Date", "Category", "Source"]
            data = ["Coq Sportif", comment, date_review, likes, datetime.today().date(), "Sportwear", "Avis-Vérifiés"]

            chemin = "D:/Lorraine/Fichiers/Coq_Sportif-Review.xlsx"

            file_exists = openpyxl.workbook.Workbook()

            file_exists = os.path.isfile(chemin)

            # Création d'un classeur s'il n'existe pas et écriture des en-têtes
            if not file_exists:
                wb = Workbook()
                ws = wb.active
                ws.append(header)
                wb.save(chemin)

            # Écriture des données
            wb = openpyxl.load_workbook(chemin)
            ws = wb.active
            ws.append(data)
            wb.save(chemin)


urls = get_all_pages()
parse_comments(urls)