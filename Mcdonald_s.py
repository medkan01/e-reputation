import datetime
import requests
from bs4 import BeautifulSoup
import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import schedule

#import schedule
import time

def get_all_pages():
    urls = []

    page_number = 1

    for i in range(75):
        i = f"https://fr.trustpilot.com/review/mcdonalds.fr?page={page_number}"
        page_number += 1
        urls.append(i)

    return urls

def parse_comments(url):
    chemin = "D:/Lorraine/Mcdonald-Review.xlsx"
    header = ['Enterprise', 'Comments', "Reviews_Date", "Evaluations", "Extraction-Date", "Category", "Source"]
    
    # Création d'un classeur s'il n'existe pas et écriture des en-têtes 
    file_exists = os.path.isfile(chemin)
    if not file_exists:
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        wb.save(chemin)
    else:
        wb = openpyxl.load_workbook(chemin)
        ws = wb.active
        # Supprimer les lignes existantes (sauf la première ligne avec les en-têtes)
        ws.delete_rows(2, ws.max_row)
        wb.save(chemin)
    
    for url in urls:
        request = requests.get(url)
        print(request.status_code)
        soup = BeautifulSoup(request.content, "html.parser")
        reviews = soup.find_all('div', class_='styles_reviewCardInner__EwDq2')
        
        for review in reviews:
            
            try:
                comment = review.find("p", class_="typography_body-l__KUYFJ typography_appearance-default__AAY17 typography_color-black__5LYEn").text.strip()
            except AttributeError as e:
                comment = ""
            
            date_review = review.find("time").get_text(strip=True)
            #print(date_review)

            likes = review.find("img")['alt']
        # print(likes)

            data = ["Mcdonald's", comment, date_review, likes, datetime.today().date(), "Fast food", "Trustpilot"]

            # Écriture des données
            wb = openpyxl.load_workbook(chemin)
            ws = wb.active
            ws.append(data)
            wb.save(chemin)


urls = get_all_pages()
#parse_comments(urls)

# Planifier le scraping toutes les 24 heures
schedule.every(15).days.do(parse_comments(urls))

# Boucle pour exécuter les tâches planifiées
while True:
    schedule.run_pending()
    time.sleep(1)
